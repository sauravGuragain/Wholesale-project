"""
Excel helpers for bulk price management.

Export produces an .xlsx of the current catalog with editable price columns.
Import reads that same shape back, validating each row and reporting per-row
errors rather than failing the whole file — partial success is surfaced to the
admin so they can fix only the bad rows.
"""
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

EXPORT_HEADERS = ["sku", "name", "cost_price", "selling_price", "tax_rate"]


@dataclass
class PriceRow:
    sku: str
    cost_price: Decimal | None
    selling_price: Decimal | None
    tax_rate: Decimal | None


@dataclass
class ParsedPriceSheet:
    rows: list[PriceRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)  # human-readable, row-numbered


def build_price_export(products: list) -> bytes:
    """products: iterable with .sku, .name, .cost_price, .selling_price, .tax_rate."""
    wb = Workbook()
    ws = wb.active
    ws.title = "prices"
    ws.append(EXPORT_HEADERS)
    for p in products:
        ws.append([p.sku, p.name, float(p.cost_price), float(p.selling_price), float(p.tax_rate)])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _to_decimal(value, *, allow_blank: bool) -> Decimal | None:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        if allow_blank:
            return None
        raise ValueError("value is required")
    try:
        d = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"'{value}' is not a number") from exc
    if d < 0:
        raise ValueError("value must be non-negative")
    return d


def parse_price_import(file_bytes: bytes) -> ParsedPriceSheet:
    result = ParsedPriceSheet()
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any openpyxl parse failure cleanly
        result.errors.append(f"Could not read the Excel file: {exc}")
        return result

    ws = wb.active
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    try:
        sku_idx = header.index("sku")
    except ValueError:
        result.errors.append("Missing required 'sku' column in header row.")
        return result

    def col(name):
        return header.index(name) if name in header else None

    cost_idx, sell_idx, tax_idx = col("cost_price"), col("selling_price"), col("tax_rate")

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        sku = row[sku_idx] if sku_idx < len(row) else None
        if not sku:
            result.errors.append(f"Row {row_num}: missing SKU.")
            continue
        try:
            cost = _to_decimal(row[cost_idx], allow_blank=True) if cost_idx is not None else None
            sell = _to_decimal(row[sell_idx], allow_blank=True) if sell_idx is not None else None
            tax = _to_decimal(row[tax_idx], allow_blank=True) if tax_idx is not None else None
        except ValueError as exc:
            result.errors.append(f"Row {row_num} (SKU {sku}): {exc}.")
            continue
        result.rows.append(PriceRow(sku=str(sku).strip(), cost_price=cost, selling_price=sell, tax_rate=tax))

    return result
