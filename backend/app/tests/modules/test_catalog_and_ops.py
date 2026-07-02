"""
Coverage for the modules added in this batch:
  * category slug generation + delete guard
  * product creation, catalog pricing/stock, Excel export/import round-trip
  * payment proof -> verify auto-confirms order
  * offers validity filtering
  * reports dashboard + best-selling
"""
from decimal import Decimal

import pytest

from app.core.exceptions import ConflictError
from app.core.security import hash_password
from app.modules.categories.schemas import CategoryCreate
from app.modules.categories.service import CategoryService
from app.modules.customers.models import Customer
from app.modules.inventory.models import Inventory
from app.modules.offers.schemas import OfferCreate
from app.modules.offers.service import OfferService
from app.modules.orders.schemas import OrderCreate, OrderItemInput
from app.modules.orders.service import OrderService
from app.modules.payments.service import PaymentService
from app.modules.products.schemas import ProductCreate
from app.modules.products.service import ProductService
from app.modules.reports.service import ReportService
from app.modules.users.models import Role, User
from app.shared.enums import DiscountType, OfferAppliesTo, OrderStatus, PaymentMethod, PaymentStatus


def _customer(db):
    role = db.query(Role).filter_by(name="customer").one()
    user = User(username="acme", password_hash=hash_password("password123"), role_id=role.id)
    db.add(user)
    db.flush()
    c = Customer(user_id=user.id, business_name="Acme Traders")
    db.add(c)
    db.commit()
    return c


def _category(db):
    return CategoryService(db).create(CategoryCreate(name="Biscuits & Snacks"))


# ---------- categories ----------

def test_category_slug_is_generated_and_unique(db_session):
    a = CategoryService(db_session).create(CategoryCreate(name="Cooking Oil"))
    b = CategoryService(db_session).create(CategoryCreate(name="Cooking Oil"))
    assert a.slug == "cooking-oil"
    assert b.slug == "cooking-oil-2"  # de-duplicated


def test_category_delete_guard_with_children(db_session):
    from app.core.exceptions import ConflictError

    parent = CategoryService(db_session).create(CategoryCreate(name="Beverages"))
    CategoryService(db_session).create(CategoryCreate(name="Juices", parent_id=parent.id))
    with pytest.raises(ConflictError):
        CategoryService(db_session).delete(parent.id)


# ---------- products + catalog ----------

def _make_product(db, sku="RICE-1", price="120.00", stock=30):
    cat = _category(db)
    p = ProductService(db).create(
        ProductCreate(name="Basmati Rice 5kg", sku=sku, category_id=cat.id, selling_price=Decimal(price), tax_rate=Decimal("5"))
    )
    inv = db.query(Inventory).filter_by(product_id=p.id).one()
    inv.quantity_on_hand = stock
    db.commit()
    return p


def test_duplicate_sku_rejected(db_session):
    _make_product(db_session, sku="DUP-1")
    with pytest.raises(ConflictError):
        _make_product(db_session, sku="DUP-1")


def test_catalog_shows_resolved_price_and_stock(db_session):
    customer = _customer(db_session)
    _make_product(db_session, sku="RICE-1", price="120.00", stock=7)

    items, total = ProductService(db_session).catalog(
        customer=customer, term=None, category_id=None, brand_id=None, page=1, page_size=20
    )
    assert total == 1
    item = items[0]
    assert item["price"] == Decimal("120.00")
    assert item["price_source"] == "default"
    assert item["in_stock"] is True
    assert item["quantity_available"] == 7


def test_excel_price_export_import_roundtrip(db_session):
    p = _make_product(db_session, sku="OIL-1", price="200.00", stock=10)
    svc = ProductService(db_session)

    blob = svc.export_prices()
    assert blob[:2] == b"PK"  # xlsx is a zip

    # Simulate an admin editing the selling price to 180 and re-uploading.
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    # header: sku,name,cost_price,selling_price,tax_rate
    ws.cell(row=2, column=4).value = 180.00
    out = io.BytesIO()
    wb.save(out)

    updated, skipped, errors = svc.import_prices(out.getvalue())
    assert updated == 1 and skipped == 0 and errors == []

    db_session.refresh(p)
    assert p.selling_price == Decimal("180.00")


def test_excel_import_reports_unknown_sku(db_session):
    _make_product(db_session, sku="KNOWN-1")
    from app.shared.utils.excel import build_price_export

    # Build a sheet referencing a SKU that doesn't exist.
    class _P:
        sku, name, cost_price, selling_price, tax_rate = "GHOST-9", "Ghost", Decimal("1"), Decimal("2"), Decimal("0")

    blob = build_price_export([_P()])
    updated, skipped, errors = ProductService(db_session).import_prices(blob)
    assert updated == 0 and skipped == 1
    assert any("GHOST-9" in e for e in errors)


# ---------- payments ----------

def test_verify_payment_confirms_pending_order(db_session):
    customer = _customer(db_session)
    p = _make_product(db_session, sku="SNK-1", price="50.00", stock=20)

    order = OrderService(db_session).place_order(
        customer=customer,
        payload=OrderCreate(items=[OrderItemInput(product_id=p.id, quantity=2)], payment_method=PaymentMethod.STATIC_QR),
        user_id=customer.user_id,
    )
    payment = PaymentService(db_session).get_by_order(order.id)
    assert payment.status == PaymentStatus.PENDING

    admin_role = db_session.query(Role).filter_by(name="admin").one()
    admin = User(username="admin1", password_hash=hash_password("password123"), role_id=admin_role.id)
    db_session.add(admin)
    db_session.commit()

    PaymentService(db_session).verify(payment_id=payment.id, admin_user_id=admin.id)

    refreshed = OrderService(db_session).get_order(order.id)
    assert refreshed.status == OrderStatus.CONFIRMED
    assert PaymentService(db_session).get_by_order(order.id).status == PaymentStatus.VERIFIED


# ---------- offers ----------

def test_offer_requires_target_for_product_scope(db_session):
    from app.core.exceptions import ValidationError

    with pytest.raises(ValidationError):
        OfferService(db_session).create(
            OfferCreate(
                name="10% off snacks",
                discount_type=DiscountType.PERCENT,
                discount_value=Decimal("10"),
                applies_to=OfferAppliesTo.PRODUCT,  # target_id missing
            )
        )


def test_active_offers_filter(db_session):
    OfferService(db_session).create(
        OfferCreate(
            name="Order discount",
            discount_type=DiscountType.FLAT,
            discount_value=Decimal("50"),
            applies_to=OfferAppliesTo.ORDER,
            is_active=True,
        )
    )
    OfferService(db_session).create(
        OfferCreate(
            name="Disabled",
            discount_type=DiscountType.FLAT,
            discount_value=Decimal("50"),
            applies_to=OfferAppliesTo.ORDER,
            is_active=False,
        )
    )
    active = OfferService(db_session).list(active_only=True)
    assert len(active) == 1 and active[0].name == "Order discount"


# ---------- reports ----------

def test_dashboard_counts_and_revenue(db_session):
    customer = _customer(db_session)
    p = _make_product(db_session, sku="REP-1", price="100.00", stock=50)
    svc = OrderService(db_session)

    order = svc.place_order(
        customer=customer,
        payload=OrderCreate(items=[OrderItemInput(product_id=p.id, quantity=2)], payment_method=PaymentMethod.COD),
        user_id=customer.user_id,
    )
    # Walk the order to delivered so it counts as revenue.
    for s in (OrderStatus.CONFIRMED, OrderStatus.PACKED, OrderStatus.OUT_FOR_DELIVERY, OrderStatus.DELIVERED):
        svc.change_status(order_id=order.id, new_status=s, user_id=customer.user_id)

    summary = ReportService(db_session).dashboard()
    assert summary["todays_orders"] == 1
    assert summary["completed_orders"] == 1
    assert summary["monthly_revenue"] == Decimal("210.00")  # 200 + 5% tax
    assert summary["best_selling"][0]["units_sold"] == 2


def test_upload_proof_returns_attached_proof(db_session):
    """Regression: the upload response must include the just-added proof, not a stale empty list."""
    from app.modules.payments.service import PaymentService

    customer = _customer(db_session)
    p = _make_product(db_session, sku="PROOF-1", price="40.00", stock=10)
    order = OrderService(db_session).place_order(
        customer=customer,
        payload=OrderCreate(items=[OrderItemInput(product_id=p.id, quantity=1)], payment_method=PaymentMethod.STATIC_QR),
        user_id=customer.user_id,
    )
    png = b"\x89PNG\r\n\x1a\n" + b"0" * 32
    payment = PaymentService(db_session).upload_proof(
        order_id=order.id, customer=customer, file_bytes=png, filename="r.png", content_type="image/png"
    )
    assert len(payment.proofs) == 1
    assert payment.status == PaymentStatus.PENDING
